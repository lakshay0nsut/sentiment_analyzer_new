# Importing the libraries
from flask import Flask, render_template, request
from googletrans import Translator
import nltk
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# Initializing the Flask app
app = Flask(__name__)

# Initialize Google Translator and NLTK Sentiment Analyzer
translator = Translator()
nltk.download('vader_lexicon')
sia = SentimentIntensityAnalyzer()

# Excel file path
EXCEL_FILE_PATH = "sentiment_analysis_results.xlsx"

# Function to record sentiment results in Excel
def record_sentiment_to_excel(review, sentiment, detected_language, file_path=EXCEL_FILE_PATH):
    # Check if the file exists
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sentiment Results"
        # Add headers to the new sheet
        sheet.append(["Review", "Sentiment", "Detected Language"])

    # Append the new data (review, sentiment, and detected language)
    sheet.append([review, sentiment, detected_language])

    # Set column widths based on the maximum length of each column
    column_widths = []
    for row in sheet.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                column_widths[i] = max(column_widths[i], len(str(cell)))
            else:
                column_widths.append(len(str(cell)))

    for i, column_width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = column_width + 2  # add some padding for readability

    # Save the workbook
    workbook.save(file_path)

# Defining the routes
@app.route('/')
def index():
    return render_template('index.html')

# Analyze route
@app.route('/analyze', methods=['POST'])
def analyze():
    if request.method == 'POST':
        review = request.form['review']
        new_review = review

        # Detecting the language of the input text
        detected_language = translator.detect(review).lang

        # Translating the input text to English if it's not already in English
        if detected_language != 'en':
            translation = translator.translate(review, dest='en')
            new_review = translation.text

        # Performing sentiment analysis
        sentiment_score = sia.polarity_scores(new_review)['compound']
        if sentiment_score > 0.05:
            sentiment = "Positive 😀"
        elif sentiment_score < -0.05:
            sentiment = "Negative 😞"
        else:
            sentiment = "Neutral 😐"

        # Record the sentiment result to Excel
        try:
            record_sentiment_to_excel(review, sentiment, detected_language)
        except Exception as e:
            return f"Error saving to Excel: {str(e)}", 500

        # Display the result to the user
        return render_template('result.html', review=review, sentiment=sentiment, detected_language=detected_language)

# Running the app
if __name__ == '__main__':
    app.run(debug=True)
